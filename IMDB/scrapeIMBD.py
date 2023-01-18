from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top rated Movies'
print(excel.sheetnames)  # rename sheet
sheet.append(['Movie Rank', 'Movie Name', 'Year of Release', 'IMDB Rating'])

try:
    source = requests.get('https://www.imdb.com/chart/top/')  # access website & return object
    source.raise_for_status()  # error control, it's good practice

    soup = BeautifulSoup(source.text, 'html.parser')  # for return all html source
    # print(soup) # print all html code

    movies = soup.find('tbody', class_="lister-list").find_all('tr')  # when need specific the use (_) with class

    # print(len(movies))
    for movie in movies:
        name = movie.find('td', class_="titleColumn").a.text
        # print(name)

        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
        # print(rank)

        year = movie.find('td', class_="titleColumn").span.text.strip('()')
        # print(year)

        rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
        # print(rating)

        # print(rank, name, year, rating)
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])

        # break
except Exception as e:
    print(e)

excel.save('IMDB Movie Ratings.xlsx')
