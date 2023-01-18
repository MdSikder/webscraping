import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
from bs4 import BeautifulSoup
import requests
import openpyxl


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Views Contents'
print(excel.sheetnames)  # rename sheet
sheet.append(['Title', 'Views', 'Posted'])

username = "admin@klovercloud.com"
password = "Hello@1234"
try:
    chrome_options = Options()
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument('user-data-dir=C:/chromedriver_win32/chromeprofile')
    chrome_driver = "C:/chromedriver_win32/chromedriver.exe"

    url = 'https://www.linkedin.com/search/results/people/?geoUrn=%5B%22106215326%22%5D&keywords=qa%20tester&origin=FACETED_SEARCH&sid=Xbp'
    driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)  # driver = webdriver.Chrome()
    driver.maximize_window()
    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(url)
    time.sleep(120)

    # "style-scope ytd-rich-item-renderer"
    # //*[@id="video-title-link"]
    # //*[@id="metadata-line"]/span[1]
    # //*[@id="metadata-line"]/span[2]

    # videos = driver.find_element(By.CLASS_NAME, 'style-scope ytd-rich-item-renderer')
    file = driver.find_elements(By.CLASS_NAME, 'reusable-search__entity-result-list list-style-none')  # remember to use fin_elements

    video_list = []
    for video in file:
        title = video.find_element(By.XPATH, ".//*[@id='nU0CzK80SkiUky6p1C2UaA==']/div/ul/li[1]/div/div/div[2]/div[1]/div[2]/div[2]").text  # (.) use in very important
        views = video.find_element(By.XPATH, ".//*[@id='nU0CzK80SkiUky6p1C2UaA==']/div/ul/li[1]/div/div/div[2]/div[2]/p").text  # (.) use in very important
        when = video.find_element(By.XPATH, ".//*[@id='nU0CzK80SkiUky6p1C2UaA==']/div/ul/li[1]/div/div/div[2]/div[1]/div[1]/div/span[1]/span/a").text  # (.) use in very important

        print(title, views, when)
        sheet.append([title, views, when])

except Exception as e:
    print(e)

excel.save('contents_title.xlsx')
