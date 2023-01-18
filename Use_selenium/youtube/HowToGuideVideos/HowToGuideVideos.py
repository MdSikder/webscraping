import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Views Contents'
print(excel.sheetnames)  # rename sheet
sheet.append(['Title', 'Views', 'Posted'])

try:
    url = 'https://www.youtube.com/@HowToGuideVideos/videos'
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    driver.get(url)
    time.sleep(80)

    file = driver.find_elements(By.CLASS_NAME, 'style-scope ytd-rich-item-renderer')  # remember to use fin_elements

    video_list = []
    for video in file:
        title = video.find_element(By.XPATH, ".//*[@id='video-title-link']").text  # (.) use in very important
        views = video.find_element(By.XPATH, ".//*[@id='metadata-line']/span[1]").text  # (.) use in very important
        when = video.find_element(By.XPATH, "//*[@id='metadata-line']/span[2]").text  # (.) use in very important

        print(title, views, when)
        sheet.append([title, views, when])

except Exception as e:
    print(e)

excel.save('contents_title.xlsx')
